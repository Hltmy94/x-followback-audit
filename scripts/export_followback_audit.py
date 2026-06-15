#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

USERNAME = "\u7528\u6237\u540d"
FOLLOW_BACK = "\u662f\u5426\u56de\u5173"
BLUE_VERIFIED = "\u662f\u5426\u84ddV\u6807\u8bc6"
SHEET_NAME = "\u672a\u56de\u5173\u5217\u8868"
NOT_FOLLOWED_BACK = "\u672a\u56de\u5173"
YES = "\u662f"
NO = "\u5426"
OUT_NAME = "X\u672a\u56de\u5173\u5ba1\u8ba1"


def desktop_dir() -> Path:
    home = Path.home()
    desktop = home / "Desktop"
    return desktop if desktop.exists() else home


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = dt.datetime.now().strftime("%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--out")
    args = parser.parse_args()

    with Path(args.input).open("r", encoding="utf-8") as f:
        payload = json.load(f)

    rows = payload.get("rows", []) if isinstance(payload, dict) else payload
    date_prefix = dt.date.today().strftime("%Y%m%d")
    out = Path(args.out) if args.out else desktop_dir() / f"{date_prefix}_{OUT_NAME}.xlsx"
    out = unique_output_path(out)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append([USERNAME, FOLLOW_BACK, BLUE_VERIFIED])
    for row in rows:
        handle = text(row.get("handle"))
        if handle:
            ws.append([handle, NOT_FOLLOWED_BACK, YES if row.get("blue_verified") else NO])

    style_header(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
